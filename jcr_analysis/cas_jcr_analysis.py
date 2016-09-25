#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import re
import operator

def open_sheet(file_workbook, sheet_name):
	"""load workbook, get sheet by name
	"""
	wb = load_workbook(file_workbook, read_only=True)
	ws = wb.get_sheet_by_name(sheet_name)

	return ws

def load_sheet_into_lists(worksheet):
	# read the worksheet to a list of lists
	lists = list()
	for row in worksheet.iter_rows(row_offset=0): # row_offset=1, skip the header
		lists.append([item.value if item.value else '' for item in row])	# process null value

	return lists

def filter_one_coloum(worksheet, index):
	"""filter out a specific column
	"""
	l = list()
	
	for row in worksheet.iter_rows(row_offset=1):
		if row[index].value:
			l.append(row[index].value.encode('utf8'))

	return l

def extract_column_from_lists(lists, index):
	"""extract a specific column from a list of lists
	"""
	return map(operator.itemgetter(index), lists)


def print_formatted_lists(lists, selected_columns):
	"""format print into fixed width (the maximum length of each column)
	"""
	# calculate the maximum length of each column
	# column_max_len = [max(len(item) for item in t) for t in zip(*lists)]
	column_max_len = [max(sum([1 if ord(c)<128 else 2 for c in item.strip()]) for item in t) for t in zip(*lists)]

	idx_jcr_partition = lists[0].index(u'中科院JCR大类分区')
	sorted_lists = [lists[0]] + sorted(lists[1:], key=operator.itemgetter(idx_jcr_partition))

	for row in sorted_lists:
		# OR '%-0*s' % (column_max_len[idx], item)

		# replacement_field ::=  "{" [field_name] ["!" conversion] [":" format_spec] "}"
		'''
		for idx, item in enumerate(row[1:], 1):
			print(idx, item, type(item))

			print('{value:<{width}}'.format(value=item.encode('utf-8'), width=column_max_len[idx]))

		'''

		s = '\t'.join(['{value:{width}}'.format(value=row[idx].strip().encode('utf-8'), width=column_max_len[idx]) 
		#s = u'\t'.join([u'{value:<{width}}'.format(value=row[idx].strip(), width=column_max_len[idx]) 
						for idx in selected_columns]) # discard the first column 'Rank'

		print(s)


def main():
	# Step 1: Load data into memory
	file_workbook = u'traces/JCR2015影响因子（所有期刊从高到低排序）+中科院分区.xlsx'
	sheet_name = u'2015年JCR'
	ws = open_sheet(file_workbook, sheet_name)

	# Step 2: Analyze
	# read the worksheet to a list of lists
	lists = load_sheet_into_lists(ws)
	header = lists[0]	# Rank	ISSN	Abbreviated Journal Title	Full Title	Category	Subcategory	Country	total Cites	IF 2014-2015	IF 2013-2014	IF 2012-2013	IF 2011-2012	IF 2010-2011	IF 2009-2010	IF 2008-2009	IF 2007-2008	5-Year Impact Factor	中科院JCR大类分区	IF升降	Immediacy Index	Articles	Cited Half-Life	Eigenfactor Score	Article Influence Score	JCR-2014收录	中科院JCR大类学科	SCI收录	本土期刊 包含港澳

	idx_category = header.index('Category')	
	idx_subcategory = header.index('Subcategory')

	'''
	all_categories = set(filter_one_coloum(ws, idx_category))
	print(len(all_categories))
	'''

	all_categories = set(extract_column_from_lists(lists, idx_category)[1:])
	print('Number of catgories', len(all_categories))

	# filter out computer science
	## based on Subcategory
	'''
	all_subcategories = set(filter_one_coloum(ws, idx_subcategory))
	cs_related_subcategories = [item for item in all_subcategories 
									if re.search('computer science|telecommunication|计算机|电信', item, re.IGNORECASE)]
	'''
	# print('\t'.join(cs_related_subcategories))
	
	'''
	计算机：软件工程	COMPUTER SCIENCE, SOFTWARE ENGINEERING
	计算机：控制论		COMPUTER SCIENCE, CYBERNETICS	
	计算机：硬件		COMPUTER SCIENCE, HARDWARE & ARCHITECTURE
	计算机：理论方法	COMPUTER SCIENCE, THEORY & METHODS			
	计算机：人工智能	#COMPUTER SCIENCE, ARTIFICIAL INTELLIGENCE
	计算机：信息系统	COMPUTER SCIENCE, INFORMATION SYSTEMS	
	计算机：跨学科应用	COMPUTER SCIENCE, INTERDISCIPLINARY APPLICATIONS
	电信学			TELECOMMUNICATIONS	
	'''

	# based on 中科院JCR大类学科
	idx_jcr_category = header.index(u'中科院JCR大类学科')
	all_jcr_categories = set(filter_one_coloum(ws, idx_jcr_category))
	cs_related_jcr_categories = [item for item in all_jcr_categories 
									if re.search('computer science|telecommunication|计算机|电信', item, re.IGNORECASE)]

	print('\t'.join(cs_related_jcr_categories)) # 计算机：软件工程	计算机：控制论	计算机：硬件	计算机：理论方法	电信学	计算机：人工智能	计算机：信息系统	计算机：跨学科应用

	# Part 3: Filter out computer science and telecommunications

	idx_country = header.index('Country')
	idx_jcr_partition = header.index(u'中科院JCR大类分区')
	indexes = range(len(header))

	#selected_columns = indexes[1:idx_category] + indexes[idx_country:idx_jcr_category] + indexes[idx_jcr_category:]
	selected_columns = [indexes[idx_jcr_category], indexes[idx_jcr_partition]] + \
						indexes[1:idx_category] + \
						indexes[idx_country:idx_jcr_partition] + \
						indexes[idx_jcr_partition+1 : idx_jcr_category] + \
						indexes[idx_jcr_category+1 :]
	
	'''
	print(selected_columns)
	for idx in selected_columns:
		print(idx)
		print(header[idx])
	'''

	# Pretty print
	#selected_jcr_category = u'计算机：跨学科应用'
	for selected_jcr_category in cs_related_jcr_categories:
		#print(selected_jcr_category, type(selected_jcr_category))

		sublists = [header] + [row for row in lists[1:] if row[idx_jcr_category]==unicode(selected_jcr_category, 'utf-8')]
		print_formatted_lists(sublists, selected_columns)

	
if __name__ == '__main__':
	main()